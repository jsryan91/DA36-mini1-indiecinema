from theater.theater_entity import *
import random
import datetime
import openpyxl as op
from movie.movie_repo import *

class TheaterRepo:
    def __init__(self):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        self.path = os.path.join(current_dir, 'theater.xlsx')
        self.date = datetime.datetime.now().strftime('%y%m%d') # excel-sheet를 date로 만들거임 !!
        self.title_time_list = []
        #---MovieRepo에서 movie_list 가지고 오기-----------#
        self.movie_repo=MovieRepo()
        self.movie_info_list=self.movie_repo.get_movie_list()
        self.movie_title_list=[self.movie_info_list[i][0] for i in range(len(self.movie_info_list))]
        #--------------------------------------------------#
        #---TheaterEntity에서 상영시간(time_list), 좌석(seat_list) 가지고 오기---#
        self.theater_entity=Theater()
        self.all_time_list=self.theater_entity.get_time_list()
        self.all_seat_list = self.theater_entity.get_seat_list()
        #--------------------------------------------------#
        self.open_xlsx()
#---------------------------------------------------------------------------------------------------#
 ## .xlsx wb 만들고 sheet 생성
    def open_xlsx(self):
        # workbook 있나 확인하기 --------------------#
        try:
            self.wb = op.load_workbook(self.path)
        except:
            self.wb = op.Workbook()
            self.wb.save(self.path)
        sheet=str(self.date)
        #--------------------------------------------#

        # worsheet 있나 확인하기 --------------------# > 없으면 만들고 movie time list도 excel에 추가하기
        if sheet not in self.wb.sheetnames:
            self.wb.create_sheet(title=sheet)
            self.wb.save(self.path)
            self.ws=self.wb[sheet]
            self.title_time_list = [[time, random.choice(self.movie_title_list)] for time in self.all_time_list]

            for i in range(len(self.all_time_list)):
                seat = self.all_seat_list[i]
                cell=self.ws.cell(row=1+i*len(seat),column=1)
                cell.value="/".join(map(str,self.title_time_list[i]))

                for j in range(len(seat)):
                    for p in range(len(seat[0])):
                        cell=self.ws.cell(row=i*len(seat)+j+1,column=p+2)
                        cell.value=seat[j][p]
        #--------------------------------------------------------------------------------------------------#

        else:
            for i in range(len(self.all_time_list)):
                self.ws=self.wb[sheet]
                cell=self.ws.cell(row=1+i*len(self.all_seat_list[i][1]),column=1)
                if cell.value is not None:
                    self.title_time_list.append(list(cell.value.split("/")))
                seat=self.all_seat_list[i]
                for j in range(len(seat)):
                    for p in range(len(seat[0])):
                        cell = self.ws.cell(row=1 + j + len(seat) * i, column=p + 2)
                        seat[j][p] = cell.value
        self.wb.save(self.path)

#---------------------------------------------------------------------------------------------------#

    def get_title_time_list(self): # 영화 및 상영시간 return
        return self.title_time_list

#---------------------------------------------------------------------------------------------------#

    def get_movie_seat_list(self,time_choice): # 영화 상영시간 별 좌석 return
        return self.all_seat_list[time_choice]

#---------------------------------------------------------------------------------------------------#

    def is_seat_empty(self, x, y,time_choice): # 고른자리가 구매가능한지 확인하기
        seat= self.all_seat_list[time_choice]
        if seat[x][y]==0: return 1
        else: return 0

#---------------------------------------------------------------------------------------------------#

    def set_seat_sold(self,x,y,time_choice):
        seat = self.all_seat_list[time_choice]
        if seat[x][y] == 0:
            cell = self.ws.cell(row=time_choice * len(seat) + (x + 1), column=y + 2)
            cell.value = 1
        self.wb.save(self.path)

#---------------------------------------------------------------------------------------------------#

    def is_seat_full(self):
        count_seat=[0 for i in range(len(self.all_seat_list))]
        for i in range(len(self.all_seat_list)):
            seat=self.all_seat_list[i]
            for j in range(len(seat)):
                for p in range(len(seat[0])):
                    if seat[j][p]==1:
                        count_seat[i]+=1
            if count_seat[i]==len(seat)*len(seat[0]):
                 count_seat[i]=1
            else:
                count_seat[i]=0
        return count_seat

#---------------------------------------------------------------------------------------------------#
